import glob
import logging
import os
import re
import shutil
import sys
import tempfile
import time
from datetime import datetime
from pathlib import Path
from tkinter import filedialog

import customtkinter as ctk
import pandas as pd
import tkinter.messagebox as messagebox
import win32com.client as win32
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, messagebox
import psutil


# ==================================================
# APP CONFIG
# ==================================================
DEFAULT_CC = "finops@protera.com"

APP_TITLE = "FinOps App"
PRIMARY_COLOR = "#0046AD"
PRIMARY_HOVER = "#003A8F"
CARD_DEFAULT = "#2B2B2B"
CARD_SELECTED = "#0046AD"
CARD_HOVER = "#1F3F7A"

EMAIL_RE = re.compile(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b")

EXCEL_BASENAME = "Accounts and Contacts"

COL_COMPANY = "Name"
COL_COID = "COID"
COL_TO = "FinOps Contacts"
COL_CSM = "CSM"
COL_TAM = "TAM"

FINOPS_TEAM_EMAIL = "finops@protera.com"
ALWAYS_CC_INTERNAL = "finops@protera.com"

FINAL_REPORTS_FOLDERS = ["Final Reports", "Final Report"]
DISPLAY_DRAFTS = False
ATTACH_ONLY_LATEST = False
FORCE_SHEET_NAME = None

FOLDER_TO_EXCEL_MAP = {
    "Airnov (Colorcon Functional Packaging)": "Airnov",
    "RouteOne": "Route One",
    "AM General Corp": "AM General",
    "Gulf States Toyota": "Gulf States Toyota",
    "Mitsubishi Motors North America": "Mitsubishi",
}

KEY_FOODS_COMPANY = "Key Food Stores Co-Operative, Inc"
KEY_FOODS_EXTRA_FOLDER = "Monthly Breakdowns"

SHURE_COMPANY = "Shure"
SHURE_RECS_FOLDER = "Recomendations"
SHURE_RECS_PREFIX = "Recommendations"

AIRNOV_COMPANY = "Airnov (Colorcon Functional Packaging)"
AIRNOV_EXTRA_FOLDER = "Final Report"

SKIP_FOLDERS = {
    "1. Reservations Expirations",
    "2. BI Dashboards",
    "3.Old Files",
    "4. Cost By server",
    "5. onboarding guides",
    "6.TESTING THE SHARING",
    "7.BUDGETS",
    "8.CloudVantage",
    "9.SD-SERVICE AREA",
    "ABIOMED",
    "Chromaflo",
    "BATON",
    "Baton",
    "Beauty Counter",
    "Direct Relief",
    "DNOW",
    "Energizer Holdings",
    "FFF Enterprises",
    "GCP Applied Technologies",
    "GE Healthcare",
    "Generac",
    "Greenheck Fan",
    "Johnsonville",
    "Milliken",
    "Old customers",
    "Enable injections",
    "K-Swiss",
    "Midwest Tapes",
    "Sothebys",
    "Sports edeavors",
    "Tableu",
    "Tidewater Inc",
    "Torex",
    "ENRU",
    "Forecasting",
    "Managecore",
    "Protera",
    "Ingevity Corporation",
    "Proterial America",
    "Rush Enterprises",
    "Kyocera",
}


# ==================================================
# LOGGING
# ==================================================
def get_base_dir() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


BASE_DIR = get_base_dir()
LOG_PATH = os.path.join(BASE_DIR, "finops_app.log")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(LOG_PATH, encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)


# ==================================================
# GENERIC HELPERS
# ==================================================
def get_current_ram_mb() -> float:
    process = psutil.Process(os.getpid())
    return process.memory_info().rss / (1024 * 1024)


def enforce_ram_limit(progress_window=None):
    current_mb = get_current_ram_mb()

    if progress_window is not None:
        progress_window.set_memory(current_mb)
    
def normalize_col(col_name: str) -> str:
    return str(col_name).strip().lower().replace("\n", " ").replace("  ", " ")


def html_escape(text: str) -> str:
    text = str(text)
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def clean_email_list(value) -> str:
    if pd.isna(value):
        return ""

    text = str(value).strip()
    if not text:
        return ""

    parts = re.split(r"[;,]", text)
    cleaned = []

    for part in parts:
        email = part.strip()
        if email and email not in cleaned:
            cleaned.append(email)

    return "; ".join(cleaned)


def combine_cc(default_cc: str, csm_value, tam_value) -> str:
    emails = []

    if default_cc:
        emails.append(default_cc.strip())

    for val in [csm_value, tam_value]:
        if pd.notna(val):
            for item in re.split(r"[;,]", str(val)):
                email = item.strip()
                if email and email not in emails:
                    emails.append(email)

    return "; ".join(emails)


def parse_sheet_date(sheet_name: str):
    try:
        return datetime.strptime(sheet_name.strip(), "%B %Y")
    except ValueError:
        return None


def get_latest_month_sheet(excel_path: str) -> str:
    xl = pd.ExcelFile(excel_path)
    valid_sheets = []

    for sheet in xl.sheet_names:
        parsed_date = parse_sheet_date(sheet)
        if parsed_date:
            valid_sheets.append((sheet, parsed_date))

    if not valid_sheets:
        raise ValueError("No worksheet was found with a format like 'February 2026'.")

    valid_sheets.sort(key=lambda x: x[1], reverse=True)
    return valid_sheets[0][0]


def load_latest_sheet_dataframe(excel_path: str):
    latest_sheet = get_latest_month_sheet(excel_path)
    df_raw = pd.read_excel(excel_path, sheet_name=latest_sheet)

    original_columns = list(df_raw.columns)
    normalized_columns = [normalize_col(col) for col in df_raw.columns]

    df = df_raw.copy()
    df.columns = normalized_columns
    df = df.fillna("")

    required_cols = ["name", "coid", "finops contacts", "csm", "tam"]
    missing = [col for col in required_cols if col not in df.columns]

    if missing:
        raise ValueError(
            f"Missing required columns: {missing}\n"
            f"Columns found: {original_columns}"
        )

    column_mapping = dict(zip(normalized_columns, original_columns))
    return df, latest_sheet, column_mapping


def create_display_value(name: str, coid: str) -> str:
    name = str(name).strip()
    coid = str(coid).strip()

    if name and coid:
        return f"{name} ({coid})"
    if name:
        return name
    return coid


def build_blank_email_html():
    return "<html><body><p><br></p></body></html>"


def build_reservation_template_html(expiration_date: str) -> str:
    safe_date = html_escape(expiration_date.strip().upper())

    return f"""
    <html>
        <body>
            <p>Hello All,</p>

            <p>
                This is to inform you that you have one (1) Savings Plan scheduled to expire on
                <b>{safe_date}</b>.
            </p>

            <p>
                Please take a moment to review the available options. You may select either a
                1-year or 3-year commitment term. All corresponding monthly charges are detailed
                in the table below.
            </p>

            <p>
                Please review your options and inform us of the servers you would like to cover,
                so we can proceed with the purchase and continue to receive the discount.
            </p>

            <p>
                For any concerns, don’t hesitate to reach out to the FinOps Team.
            </p>
        </body>
    </html>
    """


def create_outlook_draft_with_signature(
    to_emails: str,
    cc_emails: str,
    subject_text: str = "",
    body_html: str = ""
):
    outlook = win32.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)

    mail.Display()
    signature_html = mail.HTMLBody

    mail.To = to_emails
    mail.CC = cc_emails
    mail.Subject = subject_text

    if body_html.strip():
        mail.HTMLBody = body_html + signature_html
    else:
        mail.HTMLBody = build_blank_email_html() + signature_html

    return mail


def find_accounts_contacts_file():
    candidate_roots = []

    one_drive = os.environ.get("OneDrive")
    one_drive_commercial = os.environ.get("OneDriveCommercial")
    user_profile = os.environ.get("USERPROFILE")

    for root in [one_drive, one_drive_commercial, user_profile]:
        if root and os.path.exists(root):
            candidate_roots.append(root)

    patterns = [
        "Accounts and Contacts.xlsx",
        "accounts and contacts.xlsx",
        "Accounts & Contacts.xlsx",
        "accounts & contacts.xlsx",
    ]

    extra_folders = [
        "",
        "OneDrive - Protera",
        "OneDrive - Protera\\Cost Governance - Customer Specific",
        "Cost Governance - Customer Specific",
    ]

    for base_root in candidate_roots:
        for extra in extra_folders:
            search_root = os.path.join(base_root, extra) if extra else base_root
            if not os.path.exists(search_root):
                continue

            for pattern in patterns:
                direct_path = os.path.join(search_root, pattern)
                if os.path.exists(direct_path):
                    return direct_path

            try:
                for path in Path(search_root).rglob("*.xlsx"):
                    if path.name.lower() in [p.lower() for p in patterns]:
                        return str(path)
            except Exception:
                pass

    return None


# ==================================================
# MONTHLY REPORT HELPERS
# ==================================================
def resolve_root_path() -> str:
    one_drive = os.environ.get("OneDriveCommercial") or os.environ.get("OneDrive")
    if one_drive:
        candidate = os.path.join(one_drive, "Cost Governance - Customer Specific")
        if os.path.isdir(candidate):
            return candidate

    user_profile = os.environ.get("USERPROFILE")
    if user_profile:
        candidate = os.path.join(
            user_profile,
            "OneDrive - Protera",
            "Cost Governance - Customer Specific",
        )
        if os.path.isdir(candidate):
            return candidate

    raise FileNotFoundError(
        "Could not locate the synced SharePoint folder.\n\n"
        "Expected something like:\n"
        "C:/Users/<username>/OneDrive - Protera/Cost Governance - Customer Specific"
    )


def find_excel_file(root: str, basename: str) -> str:
    for ext in ("xlsx", "xlsm", "xls"):
        p = os.path.join(root, f"{basename}.{ext}")
        if os.path.exists(p):
            return p
    raise FileNotFoundError(f"Could not find '{basename}.xlsx/.xlsm/.xls' in: {root}")


def normalize_sheet_name(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower())


def get_target_sheet_name_openpyxl(excel_path: str, token: str) -> str:
    wb = load_workbook(excel_path, read_only=False, data_only=True)
    available_sheets = wb.sheetnames

    token_norm = normalize_sheet_name(token)
    for sheet in available_sheets:
        if normalize_sheet_name(sheet) == token_norm:
            return sheet

    raise ValueError(f"Could not find sheet '{token}'. Available sheets: {available_sheets}")


def extract_cell_text_and_hyperlink(cell) -> str:
    parts = []

    value = cell.value
    if value is not None:
        parts.append(str(value).strip())

    try:
        if cell.hyperlink and cell.hyperlink.target:
            target = str(cell.hyperlink.target).strip()
            if target:
                parts.append(target)
    except Exception:
        pass

    combined = " | ".join([p for p in parts if p])
    return combined.strip()


def read_contacts_sheet_with_hyperlinks(excel_path: str, sheet_name: str) -> pd.DataFrame:
    wb = load_workbook(excel_path, read_only=False, data_only=True)
    ws = wb[sheet_name]

    rows = list(ws.iter_rows())
    if not rows:
        return pd.DataFrame()

    headers = [str(c.value).strip() if c.value is not None else "" for c in rows[0]]
    data = []

    for row in rows[1:]:
        row_dict = {}
        empty_row = True

        for idx, cell in enumerate(row):
            header = headers[idx] if idx < len(headers) else f"Column_{idx + 1}"
            cell_text = extract_cell_text_and_hyperlink(cell)

            if cell_text:
                empty_row = False

            row_dict[header] = cell_text

        if not empty_row:
            data.append(row_dict)

    return pd.DataFrame(data)


def read_excel_safely(excel_path: str, token: str) -> pd.DataFrame:
    try:
        sheet_name = get_target_sheet_name_openpyxl(excel_path, token)
        return read_contacts_sheet_with_hyperlinks(excel_path, sheet_name)
    except Exception:
        pass

    tmp_dir = tempfile.mkdtemp(prefix="finopsapp_")
    local_copy = os.path.join(tmp_dir, os.path.basename(excel_path))

    last_err = None
    for _ in range(5):
        try:
            shutil.copy2(excel_path, local_copy)
            sheet_name = get_target_sheet_name_openpyxl(local_copy, token)
            return read_contacts_sheet_with_hyperlinks(local_copy, sheet_name)
        except Exception as e:
            last_err = e
            time.sleep(1)

    raise RuntimeError(f"Failed to read Excel even after local copy fallback. Last error: {last_err}")


def norm_key(s: str, is_folder: bool = False) -> str:
    s = str(s or "").strip().lower()

    if is_folder and "(" in s:
        s = s.split("(")[0].strip()

    s = re.sub(r"[\(\)\[\]\{\}]", " ", s)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def map_folder_to_excel_name(folder_name: str) -> str:
    return FOLDER_TO_EXCEL_MAP.get(folder_name, folder_name)


def prev_month_token() -> str:
    return (datetime.now() - relativedelta(months=1)).strftime("%B %Y")


def extract_emails_any_format(value) -> list[str]:
    if value is None:
        return []
    if isinstance(value, (int, float)) and pd.isna(value):
        return []

    s = str(value).strip()
    if not s or s.lower() in ["nan", "none", "null"]:
        return []

    found = EMAIL_RE.findall(s)
    out, seen = [], set()

    for e in found:
        el = e.lower()
        if el not in seen:
            seen.add(el)
            out.append(e)

    return out


def unique_preserve_order(values: list[str]) -> list[str]:
    out = []
    seen = set()

    for v in values:
        t = str(v).strip()
        if not t or t.lower() in ["nan", "none", "null"]:
            continue

        tl = t.lower()
        if tl not in seen:
            seen.add(tl)
            out.append(t)

    return out


def join_recipients(values: list[str]) -> str:
    return ";".join(unique_preserve_order(values))


def build_recipients_map(df: pd.DataFrame) -> dict[str, dict[str, str]]:
    missing_cols = [c for c in (COL_COMPANY, COL_TO, COL_CSM, COL_TAM) if c not in df.columns]
    if missing_cols:
        raise KeyError(f"Missing columns {missing_cols} in Excel. Found: {list(df.columns)}")

    recipients: dict[str, dict[str, str]] = {}

    for _, row in df.iterrows():
        company = row.get(COL_COMPANY)
        if pd.isna(company) or not str(company).strip():
            continue

        company_str = str(company).strip()

        raw_to = row.get(COL_TO)
        raw_csm = row.get(COL_CSM)
        raw_tam = row.get(COL_TAM)

        to_emails = unique_preserve_order(extract_emails_any_format(raw_to))
        csm_emails = unique_preserve_order(extract_emails_any_format(raw_csm))
        tam_emails = unique_preserve_order(extract_emails_any_format(raw_tam))
        internal_cc = unique_preserve_order(
            extract_emails_any_format(ALWAYS_CC_INTERNAL) or [ALWAYS_CC_INTERNAL]
        )

        to_list = to_emails.copy()
        cc_list = unique_preserve_order(csm_emails + tam_emails + internal_cc)

        if not to_list:
            fallback_pool = unique_preserve_order(csm_emails + tam_emails)
            if fallback_pool:
                to_list = [fallback_pool[0]]
            else:
                continue

        to_lower = {e.lower() for e in to_list}
        cc_list = [e for e in cc_list if e.lower() not in to_lower]
        cc_list = unique_preserve_order(cc_list)

        key = norm_key(company_str)
        recipients[key] = {
            "to": join_recipients(to_list),
            "cc": join_recipients(cc_list),
            "original_name": company_str,
        }

    return recipients


def find_matching_reports(company_folder: str, token: str) -> list[str]:
    patterns = [
        f"*FinOps*Review*{token}*",
        f"*Monthly*FinOps*Review*{token}*",
        f"*FinOps Review {token}*",
        f"*FinOps Review - {token}*",
        f"*FinOps Review-{token}*",
        f"*FinOps*{token}*",
    ]

    files: list[str] = []
    for folder in FINAL_REPORTS_FOLDERS:
        path = os.path.join(company_folder, folder)
        if not os.path.isdir(path):
            continue
        for pat in patterns:
            files.extend(glob.glob(os.path.join(path, pat)))

    files = list({f for f in files if os.path.isfile(f)})
    files.sort(key=os.path.getmtime, reverse=True)

    if ATTACH_ONLY_LATEST and files:
        return [files[0]]
    return files


def find_key_foods_extra_attachment(company_folder: str, token: str) -> list[str]:
    breakdown_path = os.path.join(company_folder, KEY_FOODS_EXTRA_FOLDER, token)
    if not os.path.isdir(breakdown_path):
        return []

    breakdown_patterns = [
        f"{token} Breakdown.*",
        f"{token.replace(' ', '')}Breakdown.*",
        f"*{token}*Breakdown*.*",
        f"*Breakdown*{token}*.*",
    ]

    ec2_growth_patterns = [
        f"EC2 Growth {token}.*",
        f"EC2 Growth - {token}.*",
        f"EC2 Growth_{token}.*",
        f"*EC2 Growth*{token}*.*",
        f"*{token}*EC2 Growth*.*",
    ]

    breakdown_files: list[str] = []
    ec2_growth_files: list[str] = []

    for pat in breakdown_patterns:
        breakdown_files.extend(glob.glob(os.path.join(breakdown_path, pat)))

    for pat in ec2_growth_patterns:
        ec2_growth_files.extend(glob.glob(os.path.join(breakdown_path, pat)))

    breakdown_files = list({f for f in breakdown_files if os.path.isfile(f)})
    ec2_growth_files = list({f for f in ec2_growth_files if os.path.isfile(f)})

    breakdown_files.sort(key=os.path.getmtime, reverse=True)
    ec2_growth_files.sort(key=os.path.getmtime, reverse=True)

    results = []
    if breakdown_files:
        results.append(breakdown_files[0])
    if ec2_growth_files:
        results.append(ec2_growth_files[0])

    return results


def find_airnov_extra_attachment(company_folder: str, token: str) -> list[str]:
    extra_path = os.path.join(company_folder, AIRNOV_EXTRA_FOLDER)
    if not os.path.isdir(extra_path):
        return []

    ec2_growth_patterns = [
        f"EC2 Growth {token}.*",
        f"EC2 Growth - {token}.*",
        f"EC2 Growth_{token}.*",
        f"*EC2 Growth*{token}*.*",
        f"*{token}*EC2 Growth*.*",
    ]

    files: list[str] = []
    for pat in ec2_growth_patterns:
        files.extend(glob.glob(os.path.join(extra_path, pat)))

    files = list({f for f in files if os.path.isfile(f)})
    files.sort(key=os.path.getmtime, reverse=True)

    return [files[0]] if files else []


def find_shure_recommendations_attachment(company_folder: str, token: str) -> list[str]:
    recs_path = os.path.join(company_folder, SHURE_RECS_FOLDER)
    if not os.path.isdir(recs_path):
        return []

    patterns = [
        f"{SHURE_RECS_PREFIX} {token}.*",
        f"{SHURE_RECS_PREFIX}{token.replace(' ', '')}.*",
        f"*{SHURE_RECS_PREFIX}*{token}*.*",
        f"*{token}*{SHURE_RECS_PREFIX}*.*",
    ]

    files: list[str] = []
    for pat in patterns:
        files.extend(glob.glob(os.path.join(recs_path, pat)))

    files = list({f for f in files if os.path.isfile(f)})
    files.sort(key=os.path.getmtime, reverse=True)
    return [files[0]] if files else []


def create_monthly_outlook_draft(outlook, company: str, to: str, cc: str, token: str, attachments: list[str]) -> None:
    mail = outlook.CreateItem(0)

    mail.Subject = f"{token} Cloud Consumption & FinOps Monthly Report"
    mail.To = to
    mail.CC = cc

    try:
        mail.Recipients.ResolveAll()
    except Exception:
        pass

    body_html = f"""
    Dear All,<br><br>

    As part of our continuous service improvement efforts, please find attached the {token} Cloud Consumption and FinOps Report for your review.<br>
    For your convenience, below is the agenda covered in this report:<br><br>

    &bull;&nbsp;Methodology<br>
    &bull;&nbsp;Cost Overview<br>
    &bull;&nbsp;Breakdown of Accounts<br>
    &bull;&nbsp;Observations<br>
    &bull;&nbsp;Cost Reduction Opportunities<br>
    &bull;&nbsp;Our Approach<br><br>

    Every single detail contained in the attached report can also be found in CloudVantage.<br><br>

    Should you have any questions or wish to discuss any part of this report in further detail, please do not hesitate to reach out to the FinOps team at <b>{FINOPS_TEAM_EMAIL}</b>. We are here to assist you.<br><br>

    Thank you for your continued trust and collaboration.<br><br>
    """

    mail.Display()
    signature_html = mail.HTMLBody or ""
    mail.HTMLBody = body_html + signature_html

    for f in attachments:
        if os.path.exists(f):
            mail.Attachments.Add(f)

    if DISPLAY_DRAFTS:
        mail.Display()
    else:
        mail.Save()


def write_summary_txt(
    token: str,
    successful: list[tuple[str, list[str]]],
    missing_report: list[str],
    missing_recipients: list[str],
    missing_keyfoods: list[str],
    missing_shure: list[str],
    missing_airnov: list[str],
    excluded: list[str],
) -> str:
    base_dir = get_base_dir()
    safe_token = token.replace(" ", "_")
    output_path = os.path.join(base_dir, f"Monthly_Email_Summary_{safe_token}.txt")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(f"MONTHLY EMAIL EXECUTION SUMMARY - {token}\n")
        f.write("=" * 95 + "\n\n")

        f.write(f"SUCCESSFUL DRAFTS ({len(successful)})\n")
        f.write("-" * 95 + "\n")
        if successful:
            for company, files in sorted(successful, key=lambda x: x[0].lower()):
                f.write(f"{company}\n")
                for file in files:
                    f.write(f"  - {os.path.basename(file)}\n")
                f.write("\n")
        else:
            f.write("None\n\n")

        f.write(f"\nMISSING MAIN REPORT ({len(missing_report)})\n")
        f.write("-" * 95 + "\n")
        f.write("\n".join(sorted(missing_report, key=str.lower)) or "None")

        f.write(f"\n\n\nMISSING RECIPIENTS ({len(missing_recipients)})\n")
        f.write("-" * 95 + "\n")
        f.write("\n".join(sorted(missing_recipients, key=str.lower)) or "None")

        f.write(f"\n\n\nMISSING KEY FOODS BREAKDOWN ({len(missing_keyfoods)})\n")
        f.write("-" * 95 + "\n")
        f.write("\n".join(sorted(missing_keyfoods, key=str.lower)) or "None")

        f.write(f"\n\n\nMISSING SHURE RECOMMENDATIONS ({len(missing_shure)})\n")
        f.write("-" * 95 + "\n")
        f.write("\n".join(sorted(missing_shure, key=str.lower)) or "None")

        f.write(f"\n\n\nMISSING AIRNOV EC2 GROWTH ({len(missing_airnov)})\n")
        f.write("-" * 95 + "\n")
        f.write("\n".join(sorted(missing_airnov, key=str.lower)) or "None")

        f.write(f"\n\n\nEXCLUDED FOLDERS ({len(excluded)})\n")
        f.write("-" * 95 + "\n")
        f.write("\n".join(sorted(excluded, key=str.lower)) or "None")

    try:
        os.startfile(output_path)
    except Exception:
        pass

    return output_path


def run_monthly_report_automation(
    scope: str,
    selected_company: str | None = None,
    progress_callback=None,
    log_callback=None,
) -> dict:
    def set_progress(status: str, detail: str = "", progress: float | None = None):
        if progress_callback:
            progress_callback(status, detail, progress)

    def write_log(text: str):
        logger.info(text)
        if log_callback:
            log_callback(text)

    set_progress("Starting", "Resolving SharePoint path...", 0.03)
    enforce_ram_limit()

    root = resolve_root_path()
    token = prev_month_token()

    write_log(f"Processing month token: {token}")
    set_progress("Preparing", f"Target month: {token}", 0.08)
    enforce_ram_limit()

    excel_path = find_excel_file(root, EXCEL_BASENAME)
    sheet_to_use = FORCE_SHEET_NAME if FORCE_SHEET_NAME else token

    set_progress("Loading Excel", f"Reading sheet: {sheet_to_use}", 0.15)
    enforce_ram_limit()

    df = read_excel_safely(excel_path, sheet_to_use)

    write_log(f"Loaded rows from sheet '{sheet_to_use}': {len(df)}")
    write_log(f"Sheet columns: {list(df.columns)}")

    set_progress("Building recipients", f"Loaded {len(df)} rows", 0.25)
    enforce_ram_limit()

    recipients = build_recipients_map(df)

    set_progress("Opening Outlook", "Initializing Outlook...", 0.30)
    enforce_ram_limit()

    outlook = win32.Dispatch("Outlook.Application")

    successful_drafts: list[tuple[str, list[str]]] = []
    missing_report: list[str] = []
    missing_recipients_list: list[str] = []
    missing_keyfoods: list[str] = []
    missing_shure: list[str] = []
    missing_airnov: list[str] = []
    excluded_folders: list[str] = []

    selected_key = norm_key(selected_company) if selected_company else None
    company_entries = [entry for entry in os.scandir(root) if entry.is_dir()]
    total_entries = max(len(company_entries), 1)

    for idx, entry in enumerate(company_entries, start=1):
        company = entry.name

        progress = 0.30 + (idx / total_entries) * 0.60
        set_progress("Processing company folders", f"{idx}/{total_entries} - {company}", progress)
        write_log(f"Processing: {company}")
        enforce_ram_limit()

        if company in SKIP_FOLDERS or re.match(r"^\d+\.", company):
            excluded_folders.append(company)
            continue

        if scope == "Selected Company":
            if not selected_company:
                continue

            company_norm = norm_key(company)
            company_mapped_norm = norm_key(map_folder_to_excel_name(company), is_folder=True)

            if selected_key not in {company_norm, company_mapped_norm}:
                continue

        mapped_company = map_folder_to_excel_name(company)
        key = norm_key(mapped_company, is_folder=True)

        if key not in recipients:
            key_original = norm_key(company, is_folder=True)
            if key_original in recipients:
                key = key_original
            else:
                missing_recipients_list.append(company)
                write_log(f"Missing recipients: {company}")
                continue

        files = find_matching_reports(entry.path, token)

        if company == KEY_FOODS_COMPANY:
            extra = find_key_foods_extra_attachment(entry.path, token)
            if extra:
                files.extend(extra)
            else:
                missing_keyfoods.append(company)
                write_log(f"Missing Key Foods extra attachment: {company}")

        if company == SHURE_COMPANY:
            recs = find_shure_recommendations_attachment(entry.path, token)
            if recs:
                files.extend(recs)
            else:
                missing_shure.append(company)
                write_log(f"Missing Shure recommendation attachment: {company}")

        if norm_key(company) == norm_key(AIRNOV_COMPANY):
            airnov_extra = find_airnov_extra_attachment(entry.path, token)
            if airnov_extra:
                files.extend(airnov_extra)
            else:
                missing_airnov.append(company)
                write_log(f"Missing Airnov EC2 Growth attachment: {company}")

        if not files:
            missing_report.append(company)
            write_log(f"Missing main report: {company}")
            continue

        set_progress("Creating draft", f"Creating Outlook draft for {company}", progress)
        enforce_ram_limit()

        create_monthly_outlook_draft(
            outlook=outlook,
            company=company,
            to=recipients[key]["to"],
            cc=recipients[key]["cc"],
            token=token,
            attachments=files,
        )
        successful_drafts.append((company, files))
        write_log(f"Draft created: {company}")

    set_progress("Writing summary", "Saving summary file...", 0.95)
    enforce_ram_limit()

    summary_path = write_summary_txt(
        token=token,
        successful=successful_drafts,
        missing_report=missing_report,
        missing_recipients=missing_recipients_list,
        missing_keyfoods=missing_keyfoods,
        missing_shure=missing_shure,
        missing_airnov=missing_airnov,
        excluded=excluded_folders,
    )

    set_progress("Completed", f"Summary saved to: {summary_path}", 1.0)

    return {
        "token": token,
        "successful_drafts": successful_drafts,
        "missing_report": missing_report,
        "missing_recipients": missing_recipients_list,
        "missing_keyfoods": missing_keyfoods,
        "missing_shure": missing_shure,
        "missing_airnov": missing_airnov,
        "excluded_folders": excluded_folders,
        "summary_path": summary_path,
    }


# ==================================================
# UI
# ==================================================
class ProgressWindow(ctk.CTkToplevel):
    def __init__(self, parent, title="Processing..."):
        super().__init__(parent)
        self.title(title)
        self.geometry("700x360")
        self.resizable(False, False)

        self.transient(parent)
        self.grab_set()

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(3, weight=1)

        self.title_label = ctk.CTkLabel(
            self,
            text=title,
            font=("Arial", 20, "bold")
        )
        self.title_label.grid(row=0, column=0, padx=20, pady=(20, 10), sticky="w")

        self.status_label = ctk.CTkLabel(
            self,
            text="Starting...",
            font=("Arial", 14)
        )
        self.status_label.grid(row=1, column=0, padx=20, pady=(0, 8), sticky="w")

        self.detail_label = ctk.CTkLabel(
            self,
            text="Please wait...",
            font=("Arial", 12),
            text_color="#BFC9D9"
        )
        self.detail_label.grid(row=2, column=0, padx=20, pady=(0, 10), sticky="w")

        self.progress = ctk.CTkProgressBar(self)
        self.progress.grid(row=3, column=0, padx=20, pady=(0, 10), sticky="ew")
        self.progress.set(0)

        self.memory_label = ctk.CTkLabel(
            self,
            text="RAM: 0 MB",
            font=("Arial", 12)
        )
        self.memory_label.grid(row=4, column=0, padx=20, pady=(0, 10), sticky="w")

        self.log_box = ctk.CTkTextbox(self, height=140)
        self.log_box.grid(row=5, column=0, padx=20, pady=(0, 20), sticky="nsew")

        self.update()

    def set_status(self, status: str, detail: str = "", progress: float | None = None):
        self.status_label.configure(text=status)
        self.detail_label.configure(text=detail)
        if progress is not None:
            self.progress.set(max(0, min(progress, 1)))
        self.update_idletasks()
        self.update()

    def append_log(self, text: str):
        self.log_box.insert("end", text + "\n")
        self.log_box.see("end")
        self.update_idletasks()
        self.update()

    def set_memory(self, mb: float):
        self.memory_label.configure(text=f"RAM: {mb:.1f} MB")
        self.update_idletasks()
        self.update()

class FinOpsApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.title(APP_TITLE)
        self.geometry("1080x820")
        self.minsize(920, 720)

        self.df = None
        self.latest_sheet = ""
        self.column_mapping = {}
        self.filtered_options = []
        self.display_to_row = {}
        self.all_column_widgets = []
        self.selected_mode = "View Company Information"
        self.excel_path = None

        self.build_ui()
        self.load_data()

    def build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(3, weight=1)

        self.header_frame = ctk.CTkFrame(self, corner_radius=12)
        self.header_frame.grid(row=0, column=0, padx=16, pady=(16, 8), sticky="ew")
        self.header_frame.grid_columnconfigure(0, weight=1)

        self.title_label = ctk.CTkLabel(
            self.header_frame,
            text="FinOps App",
            text_color="white",
            font=("Arial", 24, "bold")
        )
        self.title_label.grid(row=0, column=0, padx=18, pady=(14, 4), sticky="w")

        self.sheet_label = ctk.CTkLabel(
            self.header_frame,
            text="Latest sheet: loading...",
            text_color="#D9E6FF",
            font=("Arial", 13)
        )
        self.sheet_label.grid(row=1, column=0, padx=18, pady=(0, 14), sticky="w")

        self.mode_frame = ctk.CTkFrame(self, corner_radius=12)
        self.mode_frame.grid(row=1, column=0, padx=16, pady=8, sticky="ew")
        self.mode_frame.grid_columnconfigure((0, 1, 2, 3), weight=1)

        self.mode_title = ctk.CTkLabel(
            self.mode_frame,
            text="Step 1 - Select Action",
            font=("Arial", 15, "bold")
        )
        self.mode_title.grid(row=0, column=0, columnspan=4, padx=18, pady=(14, 10), sticky="w")

        self.info_tile = ctk.CTkButton(
            self.mode_frame,
            text="View Company Information\nSee company details only",
            height=84,
            fg_color=CARD_SELECTED,
            hover_color=CARD_HOVER,
            font=("Arial", 14, "bold"),
            command=lambda: self.select_mode("View Company Information")
        )
        self.info_tile.grid(row=1, column=0, padx=(18, 8), pady=(0, 14), sticky="ew")

        self.blank_tile = ctk.CTkButton(
            self.mode_frame,
            text="Blank Email\nCreate an empty Outlook draft",
            height=84,
            fg_color=CARD_DEFAULT,
            hover_color=CARD_HOVER,
            font=("Arial", 14, "bold"),
            command=lambda: self.select_mode("Blank Email")
        )
        self.blank_tile.grid(row=1, column=1, padx=8, pady=(0, 14), sticky="ew")

        self.template_tile = ctk.CTkButton(
            self.mode_frame,
            text="Reservation Expiration Template\nUse the standard Savings Plan renewal email",
            height=84,
            fg_color=CARD_DEFAULT,
            hover_color=CARD_HOVER,
            font=("Arial", 14, "bold"),
            command=lambda: self.select_mode("Reservation Expiration Template")
        )
        self.template_tile.grid(row=1, column=2, padx=8, pady=(0, 14), sticky="ew")

        self.monthly_tile = ctk.CTkButton(
            self.mode_frame,
            text="Monthly FinOps Report\nCreate monthly report drafts with attachments",
            height=84,
            fg_color=CARD_DEFAULT,
            hover_color=CARD_HOVER,
            font=("Arial", 14, "bold"),
            command=lambda: self.select_mode("Monthly FinOps Report")
        )
        self.monthly_tile.grid(row=1, column=3, padx=(8, 18), pady=(0, 14), sticky="ew")

        self.company_frame = ctk.CTkFrame(self, corner_radius=12)
        self.company_frame.grid(row=2, column=0, padx=16, pady=8, sticky="ew")
        self.company_frame.grid_columnconfigure(0, weight=1)

        self.company_title = ctk.CTkLabel(
            self.company_frame,
            text="Step 2 - Search and Select Company",
            font=("Arial", 15, "bold")
        )
        self.company_title.grid(row=0, column=0, columnspan=2, padx=18, pady=(14, 8), sticky="w")

        self.search_entry = ctk.CTkEntry(
            self.company_frame,
            placeholder_text="Type company name or COID...",
            height=36
        )
        self.search_entry.grid(row=1, column=0, padx=(18, 8), pady=(0, 10), sticky="ew")
        self.search_entry.bind("<KeyRelease>", self.on_search_change)

        self.search_button = ctk.CTkButton(
            self.company_frame,
            text="Search",
            fg_color=PRIMARY_COLOR,
            hover_color=PRIMARY_HOVER,
            command=self.apply_filter,
            width=120,
            height=36
        )
        self.search_button.grid(row=1, column=1, padx=(0, 18), pady=(0, 10), sticky="e")

        self.company_dropdown = ctk.CTkComboBox(
            self.company_frame,
            values=["Loading..."],
            height=36,
            command=self.on_dropdown_select
        )
        self.company_dropdown.grid(row=2, column=0, columnspan=2, padx=18, pady=(0, 14), sticky="ew")

        self.tabview = ctk.CTkTabview(self, corner_radius=12)
        self.tabview.grid(row=3, column=0, padx=16, pady=8, sticky="nsew")
        self.tabview.add("Company Details")
        self.tabview.add("Email Preview")
        self.tabview.add("All Excel Columns")
        self.tabview.add("Monthly Report")

        self.details_tab = self.tabview.tab("Company Details")
        self.email_tab = self.tabview.tab("Email Preview")
        self.all_columns_tab = self.tabview.tab("All Excel Columns")
        self.monthly_tab = self.tabview.tab("Monthly Report")

        # Company details tab
        self.details_tab.grid_columnconfigure(1, weight=1)

        self.selected_label = ctk.CTkLabel(
            self.details_tab,
            text="Selected Company Details",
            font=("Arial", 15, "bold")
        )
        self.selected_label.grid(row=0, column=0, padx=18, pady=(14, 12), sticky="w", columnspan=2)

        self.name_title = ctk.CTkLabel(self.details_tab, text="Name:", font=("Arial", 13, "bold"))
        self.name_title.grid(row=1, column=0, padx=18, pady=5, sticky="w")
        self.name_value = ctk.CTkLabel(self.details_tab, text="-", anchor="w", justify="left")
        self.name_value.grid(row=1, column=1, padx=18, pady=5, sticky="ew")

        self.coid_title = ctk.CTkLabel(self.details_tab, text="COID:", font=("Arial", 13, "bold"))
        self.coid_title.grid(row=2, column=0, padx=18, pady=5, sticky="w")
        self.coid_value = ctk.CTkLabel(self.details_tab, text="-", anchor="w", justify="left")
        self.coid_value.grid(row=2, column=1, padx=18, pady=5, sticky="ew")

        self.to_title = ctk.CTkLabel(self.details_tab, text="To:", font=("Arial", 13, "bold"))
        self.to_title.grid(row=3, column=0, padx=18, pady=5, sticky="nw")
        self.to_value = ctk.CTkLabel(self.details_tab, text="-", anchor="w", justify="left", wraplength=680)
        self.to_value.grid(row=3, column=1, padx=18, pady=5, sticky="ew")

        self.cc_title = ctk.CTkLabel(self.details_tab, text="CC:", font=("Arial", 13, "bold"))
        self.cc_title.grid(row=4, column=0, padx=18, pady=5, sticky="nw")
        self.cc_value = ctk.CTkLabel(self.details_tab, text="-", anchor="w", justify="left", wraplength=680)
        self.cc_value.grid(row=4, column=1, padx=18, pady=5, sticky="ew")

        self.email_editor_frame = ctk.CTkFrame(self.details_tab, fg_color="transparent")
        self.email_editor_frame.grid(row=5, column=0, columnspan=2, padx=0, pady=0, sticky="ew")
        self.email_editor_frame.grid_columnconfigure(1, weight=1)

        self.date_title = ctk.CTkLabel(self.email_editor_frame, text="Expiration Date:", font=("Arial", 13, "bold"))
        self.date_entry = ctk.CTkEntry(
            self.email_editor_frame,
            height=36,
            placeholder_text="Example: June 30, 2026"
        )
        self.date_entry.bind("<KeyRelease>", lambda event: self.update_preview())

        self.subject_title = ctk.CTkLabel(self.email_editor_frame, text="Subject:", font=("Arial", 13, "bold"))
        self.subject_entry = ctk.CTkEntry(self.email_editor_frame, height=36)
        self.subject_entry.bind("<KeyRelease>", lambda event: self.sync_subject_to_preview())

        self.body_title = ctk.CTkLabel(self.email_editor_frame, text="Body Editor:", font=("Arial", 13, "bold"))
        self.body_text = ctk.CTkTextbox(self.email_editor_frame, height=150)
        self.body_text.bind("<KeyRelease>", lambda event: self.sync_body_to_preview())

        self.date_title.grid(row=0, column=0, padx=18, pady=5, sticky="w")
        self.date_entry.grid(row=0, column=1, padx=18, pady=5, sticky="ew")
        self.subject_title.grid(row=1, column=0, padx=18, pady=5, sticky="w")
        self.subject_entry.grid(row=1, column=1, padx=18, pady=5, sticky="ew")
        self.body_title.grid(row=2, column=0, padx=18, pady=5, sticky="nw")
        self.body_text.grid(row=2, column=1, padx=18, pady=5, sticky="ew")

        # Email preview tab
        self.email_tab.grid_columnconfigure(0, weight=1)
        self.email_tab.grid_rowconfigure(1, weight=1)

        self.preview_header = ctk.CTkLabel(
            self.email_tab,
            text="Email Preview",
            font=("Arial", 15, "bold")
        )
        self.preview_header.grid(row=0, column=0, padx=18, pady=(14, 10), sticky="w")

        self.preview_box = ctk.CTkTextbox(self.email_tab)
        self.preview_box.grid(row=1, column=0, padx=18, pady=(0, 18), sticky="nsew")

        # All excel columns tab
        self.all_columns_tab.grid_rowconfigure(0, weight=1)
        self.all_columns_tab.grid_columnconfigure(0, weight=1)

        self.all_columns_scroll = ctk.CTkScrollableFrame(self.all_columns_tab, corner_radius=8)
        self.all_columns_scroll.grid(row=0, column=0, padx=14, pady=14, sticky="nsew")
        self.all_columns_scroll.grid_columnconfigure(1, weight=1)

        # Monthly report tab
        self.monthly_tab.grid_columnconfigure(1, weight=1)
        self.monthly_tab.grid_rowconfigure(3, weight=1)

        self.monthly_title = ctk.CTkLabel(
            self.monthly_tab,
            text="Monthly FinOps Report Automation",
            font=("Arial", 15, "bold")
        )
        self.monthly_title.grid(row=0, column=0, columnspan=2, padx=18, pady=(14, 12), sticky="w")

        self.scope_title = ctk.CTkLabel(self.monthly_tab, text="Run Scope:", font=("Arial", 13, "bold"))
        self.scope_title.grid(row=1, column=0, padx=18, pady=5, sticky="w")

        self.scope_dropdown = ctk.CTkComboBox(
            self.monthly_tab,
            values=["Selected Company", "All Companies"],
            height=36
        )
        self.scope_dropdown.grid(row=1, column=1, padx=18, pady=5, sticky="ew")
        self.scope_dropdown.set("Selected Company")

        self.monthly_info = ctk.CTkLabel(
            self.monthly_tab,
            text="The app will create Outlook drafts with the monthly report attachments.",
            justify="left",
            anchor="w",
            wraplength=700
        )
        self.monthly_info.grid(row=2, column=0, columnspan=2, padx=18, pady=5, sticky="ew")

        self.monthly_log = ctk.CTkTextbox(self.monthly_tab)
        self.monthly_log.grid(row=3, column=0, columnspan=2, padx=18, pady=(8, 18), sticky="nsew")

        self.monthly_run_button = ctk.CTkButton(
            self.monthly_tab,
            text="Run Monthly Report",
            fg_color=PRIMARY_COLOR,
            hover_color=PRIMARY_HOVER,
            command=self.run_monthly_report_ui,
            height=40
        )
        self.monthly_run_button.grid(row=4, column=0, columnspan=2, padx=18, pady=(0, 18), sticky="ew")

        # Bottom action frame
        self.action_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.action_frame.grid(row=4, column=0, padx=16, pady=(8, 16), sticky="ew")
        self.action_frame.grid_columnconfigure((0, 1), weight=1)

        self.refresh_button = ctk.CTkButton(
            self.action_frame,
            text="Reload Excel",
            fg_color="#444444",
            hover_color="#555555",
            command=self.load_data,
            height=40
        )
        self.refresh_button.grid(row=0, column=0, padx=(0, 8), sticky="ew")

        self.create_button = ctk.CTkButton(
            self.action_frame,
            text="Create Outlook Draft",
            fg_color=PRIMARY_COLOR,
            hover_color=PRIMARY_HOVER,
            command=self.create_email,
            height=40
        )
        self.create_button.grid(row=0, column=1, padx=(8, 0), sticky="ew")

        self.status_label = ctk.CTkLabel(
            self,
            text="Ready",
            text_color="#BFC9D9",
            font=("Arial", 12)
        )
        self.status_label.grid(row=5, column=0, padx=16, pady=(0, 12), sticky="w")

        self.select_mode("View Company Information")

    def append_monthly_log(self, text: str):
        self.monthly_log.insert("end", text + "\n")
        self.monthly_log.see("end")
        self.update()

    def load_data(self):
        try:
            if not self.excel_path:
                self.excel_path = find_accounts_contacts_file()

            if not self.excel_path or not os.path.exists(self.excel_path):
                selected_file = filedialog.askopenfilename(
                    title="Select Accounts and Contacts Excel file",
                    filetypes=[("Excel files", "*.xlsx *.xlsm *.xls")]
                )

                if not selected_file:
                    raise FileNotFoundError("The Accounts and Contacts Excel file was not found automatically.")

                self.excel_path = selected_file

            self.df, self.latest_sheet, self.column_mapping = load_latest_sheet_dataframe(self.excel_path)
            self.sheet_label.configure(
                text=f"Latest sheet: {self.latest_sheet} | File: {os.path.basename(self.excel_path)}"
            )

            self.prepare_options()
            self.search_entry.delete(0, "end")
            self.apply_filter()

            self.status_label.configure(text="Excel loaded successfully.")
        except Exception as e:
            messagebox.showerror("Error", str(e))
            self.status_label.configure(text="Load failed.")

    def select_mode(self, mode: str):
        previous_mode = self.selected_mode
        self.selected_mode = mode

        self.info_tile.configure(fg_color=CARD_SELECTED if mode == "View Company Information" else CARD_DEFAULT)
        self.blank_tile.configure(fg_color=CARD_SELECTED if mode == "Blank Email" else CARD_DEFAULT)
        self.template_tile.configure(fg_color=CARD_SELECTED if mode == "Reservation Expiration Template" else CARD_DEFAULT)
        self.monthly_tile.configure(fg_color=CARD_SELECTED if mode == "Monthly FinOps Report" else CARD_DEFAULT)

        self.update_mode_visibility()

        if mode == "Blank Email" and previous_mode != "Blank Email":
            self.reset_blank_email_fields()
        else:
            self.update_preview()

    def update_mode_visibility(self):
        if self.selected_mode == "View Company Information":
            self.email_editor_frame.grid_remove()
            self.create_button.configure(state="disabled")
            self.monthly_run_button.configure(state="disabled")
            self.tabview.set("All Excel Columns")

        elif self.selected_mode == "Blank Email":
            self.email_editor_frame.grid()
            self.date_entry.configure(state="disabled")
            self.subject_entry.configure(state="normal")
            self.body_text.configure(state="normal")
            self.create_button.configure(state="normal")
            self.monthly_run_button.configure(state="disabled")
            self.tabview.set("Email Preview")

        elif self.selected_mode == "Reservation Expiration Template":
            self.email_editor_frame.grid()
            self.date_entry.configure(state="normal")
            self.subject_entry.configure(state="disabled")
            self.body_text.configure(state="disabled")
            self.create_button.configure(state="normal")
            self.monthly_run_button.configure(state="disabled")
            self.tabview.set("Email Preview")

        else:
            self.email_editor_frame.grid_remove()
            self.create_button.configure(state="disabled")
            self.monthly_run_button.configure(state="normal")
            self.tabview.set("Monthly Report")

    def reset_blank_email_fields(self):
        self.subject_entry.configure(state="normal")
        self.subject_entry.delete(0, "end")

        self.body_text.configure(state="normal")
        self.body_text.delete("1.0", "end")

        self.update_preview()

    def prepare_options(self):
        self.display_to_row = {}

        for _, row in self.df.iterrows():
            name = str(row.get("name", "")).strip()
            coid = str(row.get("coid", "")).strip()
            display_value = create_display_value(name, coid)

            if display_value and display_value not in self.display_to_row:
                self.display_to_row[display_value] = row

    def apply_filter(self):
        search_text = self.search_entry.get().strip().lower()
        all_options = list(self.display_to_row.keys())

        if not search_text:
            filtered = sorted(all_options)
        else:
            filtered = []
            for option in all_options:
                row = self.display_to_row[option]
                name = str(row.get("name", "")).strip().lower()
                coid = str(row.get("coid", "")).strip().lower()
                display = option.lower()

                if search_text in name or search_text in coid or search_text in display:
                    filtered.append(option)

            filtered = sorted(filtered)

        if not filtered:
            filtered = ["No results found"]

        self.filtered_options = filtered
        self.company_dropdown.configure(values=self.filtered_options)
        self.company_dropdown.set(self.filtered_options[0])
        self.on_dropdown_select(self.filtered_options[0])

        if self.filtered_options[0] == "No results found":
            self.status_label.configure(text="No company matches your search.")
        else:
            self.status_label.configure(text=f"{len(self.filtered_options)} company match(es) found.")

    def on_search_change(self, event=None):
        self.apply_filter()

    def on_dropdown_select(self, selected_value):
        if selected_value == "No results found":
            self.clear_details()
            return

        row = self.display_to_row.get(selected_value)
        if row is None:
            self.clear_details()
            return

        company_name = str(row.get("name", "")).strip()
        coid = str(row.get("coid", "")).strip()
        to_emails = clean_email_list(row.get("finops contacts", ""))
        cc_emails = combine_cc(DEFAULT_CC, row.get("csm", ""), row.get("tam", ""))

        self.name_value.configure(text=company_name or "-")
        self.coid_value.configure(text=coid or "-")
        self.to_value.configure(text=to_emails or "-")
        self.cc_value.configure(text=cc_emails or "-")

        self.populate_all_columns(row)

        if self.selected_mode == "Blank Email":
            self.reset_blank_email_fields()
        else:
            self.update_preview()

    def get_selected_row(self):
        selected_company = self.company_dropdown.get()
        return self.display_to_row.get(selected_company)

    def update_preview(self):
        row = self.get_selected_row()

        self.preview_box.configure(state="normal")
        self.preview_box.delete("1.0", "end")

        if row is None:
            self.preview_box.insert("1.0", "No company selected.")
            self.preview_box.configure(state="disabled")
            return

        coid = str(row.get("coid", "")).strip()
        to_emails = clean_email_list(row.get("finops contacts", ""))
        cc_emails = combine_cc(DEFAULT_CC, row.get("csm", ""), row.get("tam", ""))
        expiration_date = self.date_entry.get().strip()

        if self.selected_mode == "View Company Information":
            self.preview_box.insert(
                "1.0",
                "Information-only mode.\n\nOpen the 'All Excel Columns' tab to review the selected company."
            )
            self.preview_box.configure(state="disabled")
            return

        if self.selected_mode == "Monthly FinOps Report":
            scope = self.scope_dropdown.get()
            preview_text = (
                f"Monthly report mode\n\n"
                f"Run Scope: {scope}\n"
                f"Selected Company: {self.company_dropdown.get()}\n\n"
                f"When you click 'Run Monthly Report', Outlook drafts will be created with the monthly report attachments."
            )
            self.preview_box.insert("1.0", preview_text)
            self.preview_box.configure(state="disabled")
            return

        if self.selected_mode == "Blank Email":
            subject_text = self.subject_entry.get().strip()
            body_value = self.body_text.get("1.0", "end").strip()

            preview_text = (
                f"To: {to_emails or '-'}\n"
                f"CC: {cc_emails or '-'}\n"
                f"Subject: {subject_text if subject_text else '(blank)'}\n\n"
                f"{body_value if body_value else '(blank email body)'}"
            )

            self.preview_box.insert("1.0", preview_text)
            self.preview_box.configure(state="disabled")
            return

        subject_text = f"Reservation Expiration coming up on {expiration_date} - {coid}"
        body_preview = (
            "Hello All,\n\n"
            "This is to inform you that you have one (1) Savings Plan scheduled to expire on "
            f"{expiration_date.upper()}.\n\n"
            "Please take a moment to review the available options. You may select either a "
            "1-year or 3-year commitment term. All corresponding monthly charges are detailed "
            "in the table below.\n\n"
            "Please review your options and inform us of the servers you would like to cover, "
            "so we can proceed with the purchase and continue to receive the discount.\n\n"
            "For any concerns, don’t hesitate to reach out to the FinOps Team."
        )

        self.subject_entry.configure(state="normal")
        self.subject_entry.delete(0, "end")
        self.subject_entry.insert(0, subject_text)
        self.subject_entry.configure(state="disabled")

        self.body_text.configure(state="normal")
        self.body_text.delete("1.0", "end")
        self.body_text.insert("1.0", body_preview)
        self.body_text.configure(state="disabled")

        preview_text = (
            f"To: {to_emails or '-'}\n"
            f"CC: {cc_emails or '-'}\n"
            f"Subject: {subject_text}\n\n"
            f"{body_preview}"
        )

        self.preview_box.insert("1.0", preview_text)
        self.preview_box.configure(state="disabled")

    def sync_subject_to_preview(self):
        if self.selected_mode == "Blank Email":
            self.update_preview()

    def sync_body_to_preview(self):
        if self.selected_mode == "Blank Email":
            self.update_preview()

    def populate_all_columns(self, row):
        for widget in self.all_column_widgets:
            widget.destroy()
        self.all_column_widgets = []

        current_row = 0

        for normalized_col in self.df.columns:
            original_col = self.column_mapping.get(normalized_col, normalized_col)
            value = row.get(normalized_col, "")

            if pd.isna(value):
                display_value = ""
            else:
                display_value = str(value).strip()

            label_col = ctk.CTkLabel(
                self.all_columns_scroll,
                text=f"{original_col}:",
                font=("Arial", 13, "bold"),
                anchor="w",
                justify="left"
            )
            label_col.grid(row=current_row, column=0, padx=(10, 15), pady=6, sticky="nw")

            value_col = ctk.CTkLabel(
                self.all_columns_scroll,
                text=display_value if display_value else "-",
                anchor="w",
                justify="left",
                wraplength=700
            )
            value_col.grid(row=current_row, column=1, padx=(0, 10), pady=6, sticky="ew")

            self.all_column_widgets.extend([label_col, value_col])
            current_row += 1

    def clear_details(self):
        self.name_value.configure(text="-")
        self.coid_value.configure(text="-")
        self.to_value.configure(text="-")
        self.cc_value.configure(text="-")

        self.subject_entry.configure(state="normal")
        self.subject_entry.delete(0, "end")

        self.body_text.configure(state="normal")
        self.body_text.delete("1.0", "end")

        self.preview_box.configure(state="normal")
        self.preview_box.delete("1.0", "end")
        self.preview_box.insert("1.0", "No company selected.")
        self.preview_box.configure(state="disabled")

        for widget in self.all_column_widgets:
            widget.destroy()
        self.all_column_widgets = []

    def create_email(self):
        if self.selected_mode in ["View Company Information", "Monthly FinOps Report"]:
            return

        selected_value = self.company_dropdown.get()

        if selected_value == "No results found":
            messagebox.showwarning("Warning", "There is no company available to select.")
            return

        row = self.display_to_row.get(selected_value)
        if row is None:
            messagebox.showwarning("Warning", "The selected company could not be found.")
            return

        company_name = str(row.get("name", "")).strip()
        coid = str(row.get("coid", "")).strip()
        to_emails = clean_email_list(row.get("finops contacts", ""))
        cc_emails = combine_cc(DEFAULT_CC, row.get("csm", ""), row.get("tam", ""))

        if not to_emails:
            messagebox.showerror(
                "Error",
                f"The company '{company_name}' does not have a value in the 'FinOps Contacts' column."
            )
            return

        if self.selected_mode == "Blank Email":
            subject_text = self.subject_entry.get().strip()
            body_raw = self.body_text.get("1.0", "end").strip()

            if body_raw:
                body_html = (
                    "<html><body>"
                    + "".join(f"<p>{html_escape(line) if line.strip() else '<br>'}</p>" for line in body_raw.splitlines())
                    + "</body></html>"
                )
            else:
                body_html = build_blank_email_html()

        else:
            expiration_date = self.date_entry.get().strip()
            if not expiration_date:
                messagebox.showwarning("Warning", "Please enter an expiration date.")
                return

            subject_text = f"Reservation Expiration coming up on {expiration_date} - {coid}"
            body_html = build_reservation_template_html(expiration_date)

        try:
            create_outlook_draft_with_signature(
                to_emails=to_emails,
                cc_emails=cc_emails,
                subject_text=subject_text,
                body_html=body_html
            )
            self.status_label.configure(text=f"Draft created for: {company_name}")
            messagebox.showinfo("Success", f"The draft was created for:\n{company_name}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create Outlook draft:\n{e}")

    def run_monthly_report_ui(self):
        if self.selected_mode != "Monthly FinOps Report":
            return

        scope = self.scope_dropdown.get()
        selected_company = None

        if scope == "Selected Company":
            selected_value = self.company_dropdown.get()
            if selected_value == "No results found":
                messagebox.showwarning("Warning", "Please select a valid company first.")
                return

            row = self.display_to_row.get(selected_value)
            if row is None:
                messagebox.showwarning("Warning", "The selected company could not be found.")
                return

            selected_company = str(row.get("name", "")).strip()

        self.monthly_log.delete("1.0", "end")
        self.append_monthly_log("Starting Monthly FinOps Report automation...")
        self.append_monthly_log(f"Run Scope: {scope}")
        if selected_company:
            self.append_monthly_log(f"Selected Company: {selected_company}")

        progress_window = ProgressWindow(self, title="Monthly FinOps Report Progress")

        def ui_progress(status, detail="", progress=None):
            progress_window.set_status(status, detail, progress)
            try:
                enforce_ram_limit(progress_window)
            except MemoryError:
                raise

        def ui_log(text):
            self.append_monthly_log(text)
            progress_window.append_log(text)

        try:
            result = run_monthly_report_automation(
                scope=scope,
                selected_company=selected_company,
                progress_callback=ui_progress,
                log_callback=ui_log,
            )

            self.append_monthly_log("")
            self.append_monthly_log(f"Month token: {result['token']}")
            self.append_monthly_log(f"Successful drafts: {len(result['successful_drafts'])}")
            self.append_monthly_log(f"Missing reports: {len(result['missing_report'])}")
            self.append_monthly_log(f"Missing recipients: {len(result['missing_recipients'])}")
            self.append_monthly_log(f"Missing Key Foods: {len(result['missing_keyfoods'])}")
            self.append_monthly_log(f"Missing Shure: {len(result['missing_shure'])}")
            self.append_monthly_log(f"Missing Airnov: {len(result['missing_airnov'])}")
            self.append_monthly_log(f"Excluded folders: {len(result['excluded_folders'])}")
            self.append_monthly_log(f"Summary saved to: {result['summary_path']}")

            if result["successful_drafts"]:
                self.append_monthly_log("")
                self.append_monthly_log("Drafts created for:")
                for company, files in result["successful_drafts"]:
                    self.append_monthly_log(f"- {company}")
                    for file in files:
                        self.append_monthly_log(f"    {os.path.basename(file)}")

            self.status_label.configure(text="Monthly report automation completed.")
            progress_window.set_status("Completed", "All tasks finished successfully.", 1.0)
            progress_window.append_log("Execution completed successfully.")

            self.after(800, progress_window.destroy)

            messagebox.showinfo(
                "Success",
                f"Monthly report automation completed.\n\nSuccessful drafts: {len(result['successful_drafts'])}"
            )

        except Exception as e:
            self.append_monthly_log("")
            self.append_monthly_log(f"ERROR: {e}")
            self.status_label.configure(text="Monthly report automation failed.")
            progress_window.set_status("Error", str(e), 1.0)
            progress_window.append_log(f"ERROR: {e}")
            messagebox.showerror("Error", str(e))
            self.after(1200, progress_window.destroy)

if __name__ == "__main__":
    app = FinOpsApp()
    app.mainloop()